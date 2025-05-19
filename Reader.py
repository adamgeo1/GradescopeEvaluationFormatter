import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os

INPUT_PATH = Path(__file__).parent.absolute() / 'ERsurvey.csv'
OUTPUT_PATH = Path(__file__).parent.absolute() / 'output'

def main():
    df = pd.read_csv(INPUT_PATH)
    # Drop the description row (row 0)
    df_clean = df.iloc[1:].copy()

    # Convert AnonID to int, replace NaNs with 0
    df_clean['AnonID'] = df_clean['AnonID'].apply(lambda x: 0 if pd.isna(x) else int(x))

    # Identify question start columns
    question_starts = [col for col in df.columns if isinstance(col, str) and col.startswith('Q')]

    # Create the result dictionary
    result = {}

    for _, row in df_clean.iterrows():
        anon_id = int(row['AnonID'])
        result[anon_id] = {}

        for q_start in question_starts:
            start_idx = df.columns.get_loc(q_start)
            next_q_idx = next(
                (df.columns.get_loc(col) for col in question_starts if df.columns.get_loc(col) > start_idx),
                len(df.columns)
            )
            question_values = row.iloc[start_idx:next_q_idx]

            # Convert to list of ints (skip NaN, leave as 0 or None if needed)
            clean_values = []
            for val in question_values:
                try:
                    clean_values.append(int(float(val)))
                except (ValueError, TypeError):
                    clean_values.append(0)

            result[anon_id][q_start] = clean_values

    formatDF = pd.DataFrame({'AnonID': result.keys()})

    for q in question_starts:
        formatDF[q] = formatDF['AnonID'].apply(
            lambda anon_id: (
                int(''.join(str(i + 1) for i, val in enumerate(result[anon_id][q]) if val == 1))
                if anon_id in result and q in result[anon_id] and any(result[anon_id][q])
                else pd.NA
            )
        )

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    summary_rows = []

    for idx in range(1, 10):
        row = {'AnonID': f'Option_{idx}'}
        for q in question_starts:
            count = formatDF[q].dropna().apply(lambda s: str(idx) in str(s)).sum()
            row[q] = count
        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)

    final_df = pd.concat([formatDF, summary_df], ignore_index=True)

    final_df.to_excel(OUTPUT_PATH / 'ERsurveyFormat.xlsx', index=False)

    excel_file= OUTPUT_PATH / 'ERsurveyFormat.xlsx'
    wb = load_workbook(excel_file)
    ws = wb.active

    summary_start_row = None
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1), 1):
        if str(row[0].value).startswith('Option_'):
            summary_start_row = i
            break

    num_options = 9
    end_row = summary_start_row + num_options - 1
    start_col = 2
    end_col = ws.max_row

    for col_idx in range(start_col, end_col + 1):
        question_title = ws.cell(row=1, column=col_idx).value
        if not question_title or not question_title.startswith('Q'):
            continue

        chart = BarChart()
        chart.title = f"{question_title} Option Counts"
        chart.x_axis.title = "Option"
        chart.y_axis.title = "Count"

        data = Reference(ws, min_col=col_idx, min_row=summary_start_row, max_row=end_row)
        categories = Reference(ws, min_col=1, min_row=summary_start_row, max_row=end_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)

        chart_anchor_col = get_column_letter(9)
        chart_anchor_row = 1 + (col_idx - start_col) * 15
        ws.add_chart(chart, f"{chart_anchor_col}{chart_anchor_row}")

    wb.save(excel_file)


    '''for q in question_starts:
        indices = ''.join(formatDF[q].dropna())
        counts = Counter(indices)
        x = sorted(counts.keys())
        y = [counts[k] for k in x]

        plt.bar(x, y)
        plt.xlabel('Rubric Index')
        plt.ylabel('Count')
        plt.title(f'{q} Response Distribution')
        plt.tight_layout()
        filename = f"{q.replace(':', '_').replace(' ', '_')}.png"
        plt.savefig(OUTPUT_PATH / filename)
        plt.close()'''

if __name__ == '__main__':
    main()