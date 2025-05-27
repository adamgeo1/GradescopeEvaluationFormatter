from MidtermReader import get_midterm_data
from GradescopeReader import get_survey_data
from scipy.stats import chi2_contingency, fisher_exact, chi2
import pandas as pd
from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt

OUTPUT_FOLDER = Path(__file__).parent.absolute() / "output"
OUTPUT_FOLDER.mkdir(exist_ok=True)
OUTPUT_PATH = OUTPUT_FOLDER / "ERsurvey_midterm_comparison.xlsx"

def main():
    midterm = get_midterm_data()
    survey = get_survey_data()

    midterm_qs = {
        'Q1' : 8,
        'Q2' : 20,
        'Q3' : 12,
        'Q4' : 40,
        'Q5_6' : 20
    }

    results = []

    for i in range(2): # tests with perfect scores and perfect - 1 scores

        if i != 0:
            results.append({
                "Survey Question": "",
                "Midterm Question": "",
                "Survey Correct & MT Correct": "",
                "Survey Correct & MT Incorrect": "",
                "Survey Incorrect & MT Correct": "",
                "Survey Incorrect & MT Incorrect": "",
                "Chi-squared": "",
                "Fisher Exact": "",
                "p-value (chi-squared)": "",
                "p-value (fisher exact)": "",
                "Notes" : ""
            })

            results.append({
                "Survey Question": f"Tests allowing for {i} less than perfect score:",
                "Midterm Question": "",
                "Survey Correct & MT Correct": "",
                "Survey Correct & MT Incorrect": "",
                "Survey Incorrect & MT Correct": "",
                "Survey Incorrect & MT Incorrect": "",
                "Chi-squared": "",
                "Fisher Exact": "",
                "p-value (chi-squared)": "",
                "p-value (fisher exact)": "",
                "Notes": ""
            })

            results.append({
                "Survey Question": "",
                "Midterm Question": "",
                "Survey Correct & MT Correct": "",
                "Survey Correct & MT Incorrect": "",
                "Survey Incorrect & MT Correct": "",
                "Survey Incorrect & MT Incorrect": "",
                "Chi-squared": "",
                "Fisher Exact": "",
                "p-value (chi-squared)": "",
                "p-value (fisher exact)": "",
                "Notes": ""
            })

        for midterm_q, score in midterm_qs.items():

            rubric_index = 0

            for survey_q in survey[next(iter(survey))].keys():  # Iterate through all survey questions
                a = b = c = d = 0

                for sid in survey:
                    if sid in midterm and midterm_q in midterm[sid] and survey_q in survey[sid]:
                        if rubric_index >= len(survey[sid][survey_q]) or rubric_index >= len(midterm[sid][midterm_q][1]):
                            continue

                        survey_correct = survey[sid][survey_q][rubric_index] == 1
                        midterm_correct = midterm[sid][midterm_q][0] >= score - i

                        if survey_correct and midterm_correct:
                            a += 1
                        elif survey_correct and not midterm_correct:
                            b += 1
                        elif not survey_correct and midterm_correct:
                            c += 1
                        else:
                            d += 1

                table = [[a, b], [c, d]]

                perform_chi = not (a < 5 or b < 5 or c < 5 or d < 5)

                if perform_chi:
                    chi2_val, p_chi, dof, expected = chi2_contingency(table)

                    x_max = max(chi2_val + 5, chi2.ppf(0.999, df=dof))
                    x = np.linspace(0, x_max, 1000)
                    y = chi2.pdf(x, df=dof)

                    plt.figure(figsize=(6, 4))
                    plt.plot(x, y, label=f'dof={dof}', color='blue')

                    x_fill = np.linspace(chi2_val, x_max, 1000)
                    y_fill = chi2.pdf(x_fill, df=dof)
                    plt.fill_between(x_fill, y_fill, color='blue', alpha=0.2, label=f'p-value = {p_chi:.4f}')

                    plt.axvline(x=chi2_val, color='red', linestyle='--', label=f'χ² = {chi2_val:.2f}')
                    plt.title(f'Chi-squared Distribution for {survey_q} vs {midterm_q}')
                    plt.xlabel('Chi-squared Value')
                    plt.ylabel('Probability Density')
                    plt.legend()
                    plt.xlim(left=0)
                    plt.ylim(bottom=0, top=.05)
                    plot_output_dir = OUTPUT_FOLDER / f"Survey_{survey_q}_VS_Midterm_{midterm_q}"
                    plot_output_dir.mkdir(exist_ok=True)
                    plot_path = plot_output_dir / f"chi_squared_{survey_q}_{midterm_q}.png" if i == 0 else plot_output_dir / f"chi_squared_{survey_q}_{midterm_q}_tests_with_{i}_less_than_perfect_score.png"
                    plt.savefig(plot_path)
                    plt.close()

                fish, p_fish = fisher_exact(table)

                results.append({
                    "Survey Question": survey_q,
                    "Midterm Question": midterm_q,
                    "Survey Correct & MT Correct": a,
                    "Survey Correct & MT Incorrect": b,
                    "Survey Incorrect & MT Correct": c,
                    "Survey Incorrect & MT Incorrect": d,
                    "Chi-squared": round(chi2_val, 4) if perform_chi else "N/A",
                    "Fisher Exact": round(fish, 4),
                    "p-value (chi-squared)": round(p_chi, 4) if perform_chi else "N/A",
                    "p-value (fisher exact)": round(p_fish, 4),
                    "Notes": "" if perform_chi else "Not enough variation for chi-squared"
                })

    alternate_results = []

    for i in range(2):

        if i != 0:
            alternate_results.append({
                "Midterm Question": "",
                "MT Correct & Score >= Median": "",
                "MT Correct & Score < Median": "",
                "MT Incorrect & Score >= Median": "",
                "MT Incorrect & Score < Median": "",
                "Chi-squared": "",
                "Fisher Exact": "",
                "p-value (chi-squared)": "",
                "p-value (fisher exact)": "",
                "Notes": ""
            })

            alternate_results.append({
                "Midterm Question": f"Tests allowing for {i} less than perfect score:",
                "MT Correct & Score >= Median": "",
                "MT Correct & Score < Median": "",
                "MT Incorrect & Score >= Median": "",
                "MT Incorrect & Score < Median": "",
                "Chi-squared": "",
                "Fisher Exact": "",
                "p-value (chi-squared)": "",
                "p-value (fisher exact)": "",
                "Notes": ""
            })

            alternate_results.append({
                "Midterm Question": "",
                "MT Correct & Score >= Median": "",
                "MT Correct & Score < Median": "",
                "MT Incorrect & Score >= Median": "",
                "MT Incorrect & Score < Median": "",
                "Chi-squared": "",
                "Fisher Exact": "",
                "p-value (chi-squared)": "",
                "p-value (fisher exact)": "",
                "Notes": ""
            })

        for midterm_q, score in midterm_qs.items():
            a = b = c = d = 0

            for sid in midterm:

                if sid in midterm and midterm_q in midterm[sid] and len(midterm[sid][midterm_q][1]) > 0:
                    mt_above_median = midterm[sid]["total_score"] - midterm_qs["Q3"] >= 73 # TODO change 73 to new medians
                    mt_correct = midterm[sid][midterm_q][0] >= score - i

                    if mt_correct and mt_above_median:
                        a += 1
                    elif mt_correct and not mt_above_median:
                        b += 1
                    elif not mt_correct and mt_above_median:
                        c += 1
                    else:
                        d += 1

            table = [[a, b], [c, d]]

            perform_chi = not (a < 5 or b < 5 or c < 5 or d < 5)
            perform_fish = not (a == 0 or b == 0 or c == 0 or d == 0)

            if perform_chi:
                chi2_val, p_chi, dof, expected = chi2_contingency(table)

                x_max = max(chi2_val + 5, chi2.ppf(0.999, df=dof))
                x = np.linspace(0, x_max, 1000)
                y = chi2.pdf(x, df=dof)

                plt.figure(figsize=(6, 4))
                plt.plot(x, y, label=f'dof={dof}', color='blue')

                x_fill = np.linspace(chi2_val, x_max, 1000)
                y_fill = chi2.pdf(x_fill, df=dof)
                plt.fill_between(x_fill, y_fill, color='blue', alpha=0.2, label=f'p-value = {p_chi:.4f}')

                plt.axvline(x=chi2_val, color='red', linestyle='--', label=f'χ² = {chi2_val:.2f}')
                plt.title(f'Chi-squared Distribution for Midterm {midterm_q} vs Midterm Score >= Median')
                plt.xlabel('Chi-squared Value')
                plt.ylabel('Probability Density')
                plt.legend()
                plt.xlim(left=0)
                plt.ylim(bottom=0, top=.05)
                plot_output_dir = OUTPUT_FOLDER / f"Midterm_{midterm_q}_VS_Midterm_Score_Greater_Than_Median"
                plot_output_dir.mkdir(exist_ok=True)
                plot_path = plot_output_dir / f"chi_squared_{midterm_q}_greater_than_median.png" if i == 0 else plot_output_dir / f"chi_squared_{midterm_q}_greater_than_median_tests_with_{i}_less_than_perfect_score.png"
                plt.savefig(plot_path)
                plt.close()

            if perform_fish:
                fish, p_fish = fisher_exact(table)

            alternate_results.append({
                "Midterm Question": midterm_q,
                "MT Correct & Score >= Median": a,
                "MT Correct & Score < Median": b,
                "MT Incorrect & Score >= Median": c,
                "MT Incorrect & Score < Median": d,
                "Chi-squared": round(chi2_val, 4) if perform_chi else "N/A",
                "Fisher Exact": round(fish, 4) if perform_fish else "N/A",
                "p-value (chi-squared)": round(p_chi, 6) if perform_chi else "N/A",
                "p-value (fisher exact)": round(p_fish, 6) if perform_fish else "N/A",
                "Notes": "" if perform_chi else "Not enough variation for chi-squared"
            })

    output_df = pd.DataFrame(results)
    alternate_output_df = pd.DataFrame(alternate_results)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False)
        alternate_output_df.to_excel(writer, sheet_name="Alternate", index=False)
        worksheet = writer.sheets["Sheet1"]
        alt_ws = writer.sheets["Alternate"]

        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.styles import PatternFill
        from copy import copy

        end_col = chr(ord('A') + len(output_df.columns) - 1)
        alt_end_col = chr(ord('A') + len(alternate_output_df.columns) - 1)
        end_row = output_df.shape[0] + 1
        alt_end_row = alternate_output_df.shape[0] + 1
        table_range = f"A1:{end_col}{end_row}"
        alt_table_range = f"A1:{alt_end_col}{alt_end_row}"

        table = Table(displayName="ERsurvey_midterm_analysis", ref=table_range)
        alt_table = Table(displayName="ERsurvey_midterm_analysis_alternate", ref=alt_table_range)
        style = TableStyleInfo(
            name="TableStyleLight15",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        alt_table.tableStyleInfo = style
        worksheet.add_table(table)
        alt_ws.add_table(alt_table)

        highlight = PatternFill(start_color="32CD44", end_color="32CD44", fill_type="solid")

        for row_idx, (survey_q, p_val_chi, p_val_fish) in enumerate(zip(output_df["Survey Question"], output_df["p-value (chi-squared)"], output_df["p-value (fisher exact)"]), start=2):
            # Bold header rows
            if isinstance(survey_q, str) and survey_q.startswith("Tests allowing for"):
                for col in range(1, len(output_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col)
                    bold_font = copy(cell.font)
                    bold_font.bold = True
                    cell.font = bold_font

            # Highlight p-value < 0.05
            if isinstance(p_val_chi, (int, float)) and p_val_chi < 0.05:
                cell = worksheet.cell(row=row_idx, column=output_df.columns.get_loc("p-value (chi-squared)") + 1)
                cell.fill = highlight

            if isinstance(p_val_fish, (int, float)) and p_val_fish < 0.05:
                cell = worksheet.cell(row=row_idx, column=output_df.columns.get_loc("p-value (fisher exact)") + 1)
                cell.fill = highlight

        for row_idx, (midterm_q, p_val_chi, p_val_fish) in enumerate(zip(alternate_output_df["Midterm Question"], alternate_output_df["p-value (chi-squared)"], alternate_output_df["p-value (fisher exact)"]), start=2):
            # Bold header rows
            if isinstance(midterm_q, str) and midterm_q.startswith("Tests allowing for"):
                for col in range(1, len(alternate_output_df.columns) + 1):
                    cell = alt_ws.cell(row=row_idx, column=col)
                    bold_font = copy(cell.font)
                    bold_font.bold = True
                    cell.font = bold_font

            # Highlight p-value < 0.05
            if isinstance(p_val_chi, (int, float)) and p_val_chi < 0.05:
                cell = alt_ws.cell(row=row_idx, column=alternate_output_df.columns.get_loc("p-value (chi-squared)") + 1)
                cell.fill = highlight

            if isinstance(p_val_fish, (int, float)) and p_val_fish < 0.05:
                cell = alt_ws.cell(row=row_idx, column=alternate_output_df.columns.get_loc("p-value (fisher exact)") + 1)
                cell.fill = highlight

        for col_idx, column in enumerate(output_df.columns, start=1):
            header_text = str(column)
            adjusted_width = len(header_text) + 2  # slight padding
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = adjusted_width

        for col_idx, column in enumerate(alternate_output_df.columns, start=1):
            header_text = str(column)
            adjusted_width = len(header_text) + 2  # slight padding
            col_letter = alt_ws.cell(row=1, column=col_idx).column_letter
            alt_ws.column_dimensions[col_letter].width = adjusted_width

    print(f"\nAll results written to {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
